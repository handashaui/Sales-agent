"""Generate an Excel workbook for sales agent evaluation cases and results."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                vertical="top", wrap_text=True
            )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 48)


def load_eval_results(path: Path | None) -> dict[str, Any]:
    if not path or not path.exists():
        return {"summary": {}, "details": []}
    return json.loads(path.read_text(encoding="utf-8"))


def build_workbook(cases: list[dict[str, Any]], eval_results: dict[str, Any]) -> Workbook:
    wb = Workbook()
    ws_cases = wb.active
    ws_cases.title = "cases"
    write_sheet(
        ws_cases,
        [
            "id",
            "name",
            "lead_id",
            "conversation",
            "expected_state",
            "forbidden_terms",
            "forbidden_tools",
        ],
        [
            {
                "id": case["id"],
                "name": case["name"],
                "lead_id": case["lead_id"],
                "conversation": json.dumps(case["conversation"], ensure_ascii=False),
                "expected_state": json.dumps(case.get("expected_state", {}), ensure_ascii=False),
                "forbidden_terms": ", ".join(case.get("forbidden_terms", [])),
                "forbidden_tools": ", ".join(case.get("forbidden_tools", [])),
            }
            for case in cases
        ],
    )

    ws_tools = wb.create_sheet("expected_tools")
    write_sheet(
        ws_tools,
        ["case_id", "expected_tool_sequence"],
        [
            {
                "case_id": case["id"],
                "expected_tool_sequence": " > ".join(case.get("expected_tools", [])),
            }
            for case in cases
        ],
    )

    ws_rubric = wb.create_sheet("rubric")
    write_sheet(
        ws_rubric,
        ["dimension", "description", "scoring"],
        [
            {
                "dimension": "policy_compliance",
                "description": "Avoids budget-first dismissal and unsupported claims.",
                "scoring": "1 if compliant, else 0",
            },
            {
                "dimension": "tool_correctness",
                "description": "Expected tool sequence appears in order.",
                "scoring": "1 if ordered expected tools are present, else 0",
            },
            {
                "dimension": "no_fabrication",
                "description": "Does not invent pricing, cases, ROI, or commitments.",
                "scoring": "1 if forbidden terms absent, else 0",
            },
            {
                "dimension": "handoff_correctness",
                "description": "Triggers human handoff for explicit or risky requests.",
                "scoring": "1 if handoff state matches expected state",
            },
            {
                "dimension": "crm_update",
                "description": "Writes CRM notes when qualification, demo, or handoff happens.",
                "scoring": "1 if crm_updated matches expected state",
            },
            {
                "dimension": "demo_flow",
                "description": "Completes calendar check, booking, and CRM update for valid demo requests.",
                "scoring": "1 if demo_status matches expected state",
            },
            {
                "dimension": "json_parseable",
                "description": "Output includes stable assistant_message, tool_calls, and state.",
                "scoring": "1 if required keys are present",
            },
            {
                "dimension": "llm_as_judge",
                "description": "Optional OpenAI judge score, heuristic fallback when no API key exists.",
                "scoring": "0 to 1",
            },
        ],
    )

    ws_results = wb.create_sheet("eval_results")
    rows = eval_results.get("summary", {}).get("rows", [])
    if rows:
        write_sheet(
            ws_results,
            ["case_id", "case_name", "overall_score", "fixed_score", "llm_as_judge", "passed"],
            rows,
        )
    else:
        write_sheet(
            ws_results,
            ["case_id", "case_name", "overall_score", "fixed_score", "llm_as_judge", "passed"],
            [],
        )
    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate eval cases Excel workbook.")
    parser.add_argument("--cases", default="evals/sales_cases.json")
    parser.add_argument("--eval-results", default="eval_results/eval_results.json")
    parser.add_argument("--output", default="evals/sales_agent_eval_cases.xlsx")
    args = parser.parse_args()

    cases = json.loads(Path(args.cases).read_text(encoding="utf-8"))
    eval_results = load_eval_results(Path(args.eval_results))
    wb = build_workbook(cases, eval_results)
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
